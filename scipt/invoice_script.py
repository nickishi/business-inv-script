import tkinter as tk
import csv
import os.path
from datetime import date as date_function
from fpdf import FPDF
from openpyxl import load_workbook



def createPdf(invoice_contents):

    
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)
    pdf.cell(200, 10, txt="Testing with List Item: " + invoice_contents[0], ln=1, align="C")
    pdf.cell(300, 10, txt = "Testing with list Item 2: " + invoice_contents[1], ln = 2, align = "C")
    ###SET FILE NAME TO INVOICE NUMBER
    pdf.output("simple_demo.pdf")


def enterToCsv():
    ###ADD INFORMATION FROM LUIS MASTERLOG 1/10/23


    ###Old code for sending info to a CSV, but user wants .xlsx format###
    #if os.path.exists("Master_Log_of_Invoices.csv") == False:

    #    with open("Master_Log_of_Invoices.csv", "w", newline ='') as file:
    #        writer = csv.writer(file, lineterminator = "\n")
            
    #        writer.writerow(["F", "Billed To", "Site", "Amount", "Date", "Tag Number", "Disposal Site", "Original Weight", "Edited Weight", "Date Sent", "Invoice", "Paid Date", "Payment Number", "Notes"])


    
    inv_get = inv_num.get()
    tag_get = tag_num.get()
    #labor_get = labor_num.get()
    bill_basic_get = bill_basic_num.get()
    bill_weight_get = float(bill_weight_num.get())
    weight_get = float(weight_num.get())
    bill_to_get = str(bill_to_num.get())
    site_address_get = str(site_address_num.get())
    site_disp_get = str(site_disp_num.get())
    weight_disp_get = float(weight_disp_num.get())

    bill_weight_total = float(bill_weight_get) * float(weight_get)
    bill_total = round(float(weight_disp_get) + float(bill_weight_total), 2)

    print(inv_get)
    today_date = date_function.today()
    date_format = today_date.strftime("%B %d, %Y")


    workbook_name = "Master_Log_of_Invoices_.xlsx"
    wb = load_workbook(workbook_name)

    page = wb.active

    new_data = [inv_get, bill_to_get, site_address_get, "$" + str(bill_total), date_format, tag_get, site_disp_get, weight_get, "", "", inv_get,"", "",""]

    page.append(new_data)

    wb.save(filename = workbook_name)


    ###Old code for sending info to a CSV, but user wants .xlsx format###
    #with open("Master_Log_of_Invoices.csv", "a", newline='') as file:
    #    writer = csv.writer(file, lineterminator = "\n")
    #    #writer.writerow(["F", "Billed To", "Site", "Amount", "Date", "Tag Number", "Disposal Site", "Original Weight", "Edited Weight", "Date Sent", "Invoice", "Paid Date", "Payment Number", "Notes"])
    #    writer.writerow([inv_get, bill_to_get, site_address_get, "$" + str(bill_total), date_format, tag_get, site_disp_get, weight_get, "", "", inv_get,"", "",""])

    invoice_contents = [inv_get, tag_get, bill_basic_get]

    createPdf(invoice_contents)


def tab_order():
    widgets = [inv_num, tag_num, labor_hours_num, bill_weight_num, weight_num, weight_disp_num, bill_to_num, site_address_num]
    for widget in widgets:
        widget.lift()




def main():

    #Declaring Entry variables for class enterToCsv

    ### ADD LABELS AND ENTRY BOXES TO REFLECT LUIS MASTERLOG 1/10/23
    global inv_num
    global weight_num
    global tag_num
    #global labor_num
    global bill_basic_num
    global bill_weight_num
    global weight_disp_num
    global date
    global bill_to_num
    global site_address_num
    global labor_hours_num
    global site_disp_num

    window = tk.Tk()
    window.geometry("700x300")
    
    #Labels for app

    inv_label = tk.Label(window, text = "Invoice #:")
    weight_label = tk.Label(text = "Weight:")
    edit_weight_label = tk.Label(text = "Edited Weight: ")
    tag_label = tk.Label(text = "Tag #:")
    #labor_label = tk.Label(text = "Labor Description: ")
    #bill_basic_label = tk.Label(text = "Base Bill Rate: ")
    bill_weight_label = tk.Label(text = "Labor Fee: ")
    labor_hours_label = tk.Label(text = "Labor Hours: ")
    weight_disp_label = tk.Label(text = "Weight Disposal Fee: ")
    bill_to_label = tk.Label(text = "Bill to: ")
    site_address_label = tk.Label(text = "Site Address: ")
    site_disp_label = tk.Label(text = "Disposal Site: ")


    #labor_amount_label = tk.Label(text = "Weight of Load (in Tons): ")


    #Entry Windows
    inv_num = tk.Entry(window, width = 10)
    weight_num = tk.Entry(window, width = 10)
    tag_num = tk.Entry(window, width = 10)
    #labor_num = tk.Entry(window, width = 10)
    bill_basic_num = tk.Entry(window, width = 10)
    bill_weight_num = tk.Entry(window, width = 10)
    labor_hours_num = tk.Entry(window, width = 10)
    weight_disp_num = tk.Entry(window, width = 10)
    bill_to_num = tk.Entry(window, width = 40)
    site_address_num = tk.Entry(window, width = 40)
    site_disp_num = tk.Entry(window, width = 40)
    #labor_amount_num = tk.Entry(window, width = 3)


    #Submit Button to Create PDF and Enter into CSV
    #CREATE DEF FOR BOTH
    sub_button = tk.Button(text = "Submit", command = enterToCsv)


    #Grid Positions of Label Windows
    inv_label.grid(row = 0, column = 0)

    tag_label.grid(row = 1, column = 0, sticky = "e")

    labor_hours_label.grid(row = 2, column = 0, sticky = "e")
    
    #labor_label.grid(row = 0, column = 2, sticky = "e")
    weight_disp_label.grid(row = 2, column = 2, sticky = "e")

    #bill_basic_label.grid(row = 2, column = 2, sticky = "e")

    weight_label.grid(row = 1, column = 2, sticky = "e")

    bill_weight_label.grid(row = 0, column = 2, sticky = "e")

    bill_to_label.grid(row = 0, column = 4, sticky = "e")

    site_address_label.grid(row = 1, column = 4, sticky = "e")

    site_disp_label.grid(row = 2, column = 4, sticky = "e")
    #labor_amount_label.grid(row = 1, column = 2)


    
    #Grid Positions of Entry Windows
    inv_num.grid(row = 0, column = 1)

    tag_num.grid(row = 1, column = 1)

    labor_hours_num.grid(row = 2, column = 1)
    
    #labor_num.grid(row = 0, column = 3)

    bill_weight_num.grid(row = 2, column = 3, sticky = "w")

    weight_num.grid(row = 1, column = 3, sticky = "w")

    bill_weight_num.grid(row = 0, column = 3, sticky = "w")

    weight_disp_num.grid(row = 2, column = 3, sticky = "w")
    #labor_amount_num.grid(row = 1, column = 3)
    bill_to_num.grid(row = 0, column = 5, sticky = "w")

    site_address_num.grid(row = 1, column = 5, sticky = "w")

    site_disp_num.grid(row = 2, column = 5, sticky = "w")
   
    sub_button.grid(row = 6, column = 6)
    
    #window.title(text = "")
    
    tab_order()

    window.mainloop()


if __name__ == "__main__":
    main()
